import pandas as pd
import os
import sys
sys.path.append(r'databricks_path')
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from vehicle_telematics_codes_with_temp.matrix_creation import create_gear_matrix
from vehicle_telematics_codes_with_temp.config import result_path,Coeff_path,template_path
from vehicle_telematics_codes_with_temp.data_cleaning import process_data

def main():
    # Ensure the module path is available
    global Coeff_path
    cp=Coeff_path
    sys.path.append(r'databricks_path')

    # Define root path
    root_path = r"catalog_path"
    print("Available folders:", os.listdir(root_path))

    path,vid=get_data_folder(root_path)
    print(f"Processing data from: {path}")
    df = process_data(path,vid,cp)
    gear_matrices = create_gear_matrix(df)
    print("Gear matrices created successfully.")
    return gear_matrices,vid

def export_result(data_dict, vchl_id, mil_total, result_path, template_path,vehicle,total_rn_time):
    try:
        # Ensure result_path exists
        os.makedirs(result_path, exist_ok=True)

        # Load the template workbook
        book = load_workbook(template_path)

        # Ensure all sheets are visible before writing
        for ws in book.worksheets:
            ws.sheet_state = 'visible'

        # Define expected sheet names
        expected_sheets = {"Mileage_values", "Mileage_percentages"}

        # Validate template sheets
        template_sheets = {ws.title for ws in book.worksheets}
        missing_sheets = expected_sheets - template_sheets
        if missing_sheets:
            raise ValueError(f"Template is missing sheets: {missing_sheets}. Expected: {expected_sheets}")

        # Validate data_dict keys
        invalid_keys = [key for key in data_dict.keys() if key not in expected_sheets]
        if invalid_keys:
            raise ValueError(f"Invalid keys in data_dict: {invalid_keys}. Expected: {expected_sheets}")

        # Additional check: Ensure all expected sheets have data
        missing_data_keys = expected_sheets - data_dict.keys()
        if missing_data_keys:
            raise ValueError(f"Missing data for sheets: {missing_data_keys}. All expected sheets must have data.")

        # Define output file path
        output_file = os.path.join(result_path, f"{vchl_id}_Speed_Torque_mapping.xlsx")

        # Write NumPy arrays directly using openpyxl
        for key, matrix in data_dict.items():
            if key in template_sheets:
                ws = book[key]
                start_row, start_col = 9, 4
                for i, row in enumerate(matrix):
                    for j, val in enumerate(row):
                        ws.cell(row=start_row + i, column=start_col + j, value=val)

                # Apply color scale for Mileage_values
                
                if key == "Mileage_values":
                    max_row = start_row + matrix.shape[0] - 1
                    max_col = start_col + matrix.shape[1] - 1
                    cell_range = f"{ws.cell(row=start_row, column=start_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"

                    # Apply Red → White color scale
                    color_rule = ColorScaleRule(
                        start_type='min', start_color='FF0000',  # Red for min
                        mid_type='percentile', mid_value=50, mid_color='FFC7CE',  # Light pink mid
                        end_type='max', end_color='FFFFFF'  # White for max
                    )
                    ws.conditional_formatting.add(cell_range, color_rule)

                    # Hide zeros using custom number format
                    for i in range(start_row, start_row + matrix.shape[0]):
                        for j in range(start_col, start_col + matrix.shape[1]):
                            cell = ws.cell(row=i, column=j)
                            if cell.value == 0:
                                cell.number_format = ';;;'
                            else:
                                cell.number_format = '0.00'  # Keep numeric format

                # Apply percentage format for Mileage_percentages
                if key == "Mileage_percentages":
                    for i in range(start_row, start_row + matrix.shape[0]):
                        for j in range(start_col, start_col + matrix.shape[1]):
                            cell = ws.cell(row=i, column=j)
                            if cell.value == 0:
                                cell.number_format = ';;;'
                            else:
                                cell.number_format = '0.00%'  # Percentage format


        # Insert metadata
        for ws in book.worksheets:
            ws.sheet_state = 'visible'
            ws['B1'] = vchl_id
            ws['B2'] = vehicle
            ws['B3']=mil_total
            ws['B4']=total_rn_time


        # Save workbook
        book.save(output_file)
        print(f"New result workbook created successfully at: {output_file}")

    except Exception as e:
        print(f"Error exporting result: {e}")

def get_data_folder(path):
    import os

    if not os.path.exists(path):
        print("Path does not exist. Please enter a valid path.")
        return get_data_folder(input("Enter a valid base path: "))

    if os.path.isdir(path):
        print("Subfolders:", os.listdir(path),sep='\n')
        subpath = input("Enter your subfolder name (or press Enter to use current folder): ").strip()
        if subpath:
            new_path = os.path.join(path, subpath)
            return get_data_folder(new_path)
        else:
            return path,path.split('/')[-1]
    else:
        print("File path selected:", path)
        return path,path.split('/')[-1]

if __name__ == "__main__":
    create_gear_matrix()